import openpyxl
import re

def metabolite_bacteria_cleaning(inputfile):
    # 输入文件inputfile
    workbook = openpyxl.load_workbook(inputfile)
    worksheet = workbook['NutrientRequirementReferences']

    # row4代表inputfile所有列列名
    row4 = [item.value for item in list(worksheet.rows)[4]]

    # 输出文件outbook
    outbook = openpyxl.load_workbook('metabolite_bacteria01.xlsx')
    outsheet = outbook.create_sheet()
    new = 0

    for start in range(5,len(list(worksheet.rows))):
        # 遍历inputfile所有行row
        row = [item.value for item in list(worksheet.rows)[start]]

        # 遍历每个row的pmid,reference单元格正则匹配
        pmids = []
        # for i in range(97,len(row)):
        # for i in range(30,len(row)):
        for i in range(56,len(row)):
            pmid = re.findall(r"PMID:", str(row[i]))
            if pmid != []:
                pmids.append(str(row[i]))

        for column in range(2,56):
            # 遍历单行的所有单元格row_temp
            row_temp = row[column]
            # 单行的列头row_head
            row_head = row4[column]
            li = []
            # 找到所有值为1的单元格
            if row_temp == -1:
                li.append(row[0])
                li.append(row_head)
                li = li + pmids
                outsheet.append(li)
    outbook.save('metabolite_bacteria01.xlsx')


def bacteria_metabolism_compound_cleaning(inputfile):
    wb = openpyxl.load_workbook(inputfile)
    wsb = wb['b']
    # wsc = wb['c']
    wsd = wb['d']
    # wse = wb['e']
    wse = wb['Sheet']
    wsf = wb['f']
    new_sheet = wb.create_sheet()

    for start in range(len(list(wse.rows))):
        li = []
        rowe = [item.value for item in list(wse.rows)[start]]
        li = li + rowe
        # for brows in range(3,len(list(wsb.rows))):
        #     # rowc = [itemc.value for itemc in list(wsc.rows)[brows]]
        #     rowb = [itemc.value for itemc in list(wsb.rows)[brows]]
        #     if rowb[0] == rowe[3]:
        #         li.append(rowb[3])
        #         break
        for n in range(4,len(list(wsd.rows))):
            rowd = [itemc.value for itemc in list(wsd.rows)[n]]
            if rowd[1] == rowe[1]:
                li.append(rowd[2])
                break
        for i in range(3,len(list(wsf.rows))):
            rowf = [itemc.value for itemc in list(wsf.rows)[i]]
            if rowf[0] == rowe[1]:
                li.append(rowf[1])
                break
        new_sheet.append(li)
        if start%10 == 0:
            print(start)
    wb.save('new.xlsx')
def bacteria_reference_cleaning(inputfile):
    wb = openpyxl.load_workbook(inputfile)
    wsb = wb['b']
    # wse = wb['e']
    wse = wb['new']
    new_sheet = wb.create_sheet()

    for start in range(len(list(wse.rows))):
    # for start in range(10):
        li = []
        rowe = [item.value for item in list(wse.rows)[start]]
        li = li + rowe
        for brows in range(3,len(list(wsb.rows))):
            # rowc = [itemc.value for itemc in list(wsc.rows)[brows]]
            rowb = [itemc.value for itemc in list(wsb.rows)[brows]]
            # print(rowe, rowb)

            if str(rowb[0]) == rowe[3]:
                # print(rowb, rowe)
                li.append(rowb[3])
                break


        new_sheet.append(li)
        if start%10 == 0:
            print(start)
    wb.save('new.xlsx')










if __name__ == '__main__':
    # metabolite_bacteria_cleaning('metabolite_bacteria.xlsx')
    # bacteria_metabolism_compound_cleaning('bacteria_metabolism_compound.xlsx')
    bacteria_reference_cleaning('new.xlsx')